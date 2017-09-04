#!/usr/bin/python
from openpyxl import load_workbook
import pandas as pd
import getopt
import sys


def value_changer(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.fill.start_color.index == 'FFFF0000':
                cell.value = '<font color="red">' + cell.value + '</font>'
            if cell.font.color.index == 'FFFF0000':
                cell.value = '<font color="red">' + cell.value + '</font>'
    return worksheet


def format_changer(inputfile, outputfile):
    wb = load_workbook(filename=inputfile, data_only=True)
    ws = wb.worksheets[0]
    value_changer(ws)
    df = pd.DataFrame(ws.values)
    del ws
    del wb
    new_header = df.iloc[0]
    df = df[1:]
    df.rename(columns=new_header, inplace=True)
    df.fillna(value='', inplace=True)
    df.to_html(outputfile, index=False, header=True, escape=False)
    del df


def main(argv):
    inputfile = ''
    outputfile = ''

    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["ifile=", "ofile="])
    except getopt.GetoptError:
        print('xlsxloader.py -i <"inputfile"> -o <"outputfile">')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('xlsxloader.py -i <"inputfile"> -o <"outputfile">')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg

    if (inputfile != '') and (outputfile != ''):
        try:
            format_changer(inputfile, outputfile)
        except IOError:
            print("Inappropriate file paths.")
    else:
        print("Input and output file paths are required.")


if __name__ == "__main__":
    main(sys.argv[1:])
